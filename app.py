from flask import Flask, request, Response
from twilio.twiml.voice_response import VoiceResponse, Gather
import requests
import os
from dotenv import load_dotenv
import urllib.parse
import time
import openai
import threading
from functools import lru_cache
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime

load_dotenv()

# ─── CONFIG ────────────────────────────────────────────────────────────────────
# Exotel Configuration
account_sid = os.getenv("EXOTEL_SID") or 'getfarms1'
auth_token = os.getenv("EXOTEL_TOKEN") or 'f79262e6dfa19f80e40ed79502e172edaf4ad0baaac2a547'
exotel_subdomain = os.getenv("EXOTEL_SUBDOMAIN") or 'api.exotel.com'
exotel_number = os.getenv("EXOTEL_NUMBER") or '095-138-86363'

# ElevenLabs
ELEVEN_API_KEY = os.getenv("ELEVENLABS_API_KEY") or 'sk_d9f1be11588b6b2845b28f92e4d53bc0a06e7c923b33dc2f'
voice_id = 'd3ThhkptBvTdyy8WfbdW'

# OpenAI
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY") or '3bDJaP7gC0qqoa9tqTMprPWskG5anntGiwuPquvKdCiQkivBcqUYJQQJ99BDACHYHv6XJ3w3AAAAACOG9riH'
openai.api_key = OPENAI_API_KEY

# Excel Configuration
EXCEL_FILE = "Contact_list.xlsx"
SHEET_NAME = "Leads"
STATUS_COLORS = {
    "interested": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
    "not interested": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    "no response": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    "call failed": PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
}

# Conversation script with pre-generated audio URLs
SCRIPT = [
    {
        "text": "வணக்கம் sir, இது ஷ்ருதி கேட் ஃபார்ம்ஸிலிருந்து",
        "audio_url": None  
    },
    {
        "text": "மாந்தோப்பு விஷயமா enquiry பண்ணியிருந்தீங்க. அதுக்கான டீடெயில்ஸ் கொடுக்கதுக்காக தான் உங்களை contact பண்ணி இருக்கேன்.",
        "audio_url": None
    },
    {
        "text": "நம்முடைய மாந்தோப்பு முப்பது ஏக்கர் property. அதுல 22 சென்ட், 18 லட்ச ரூபாய்க்கு minimum ல குடுத்திட்டு இருக்கோம். இதுல நீங்க கால் ஏக்கர், அரை ஏக்கர், ஒரு ஏக்கர் purchase பண்ணிக்கலாம். இதைப் பற்றிய further information காக நம்ம team உங்களை contact பண்ணுவாங்க.",
        "audio_url": None
    }
]

# Pre-defined responses with audio caching
RESPONSES = {
    "interested": {
        "text": "நல்லது சர், நன்றி. இப்போ என் manager உங்களை call பண்ணுவாங்க.",
        "audio_url": None
    },
    "not interested": {
        "text": "சரி சர், future updates காக Getfarms ல contact பண்ணுங்க. உங்க support கும் patience கும் ரொம்ப thanks சர்.",
        "audio_url": None
    },
    "default": {
        "text": "நன்றி சர், உங்கள் நேரத்திற்கு. எங்களை தொடர்பு கொள்ளவும்.",
        "audio_url": None
    }
}

# ────────────────────────────────────────────────────────────────────────────────

app = Flask(__name__)

@lru_cache(maxsize=100)
def generate_audio(text: str) -> bytes:
    """Generate and cache Tamil speech using ElevenLabs"""
    url = f"https://api.elevenlabs.io/v1/text-to-speech/{voice_id}/stream"
    headers = {'xi-api-key': ELEVEN_API_KEY, 'Content-Type': 'application/json'}
    payload = {
        'text': text,
        'model_id': 'eleven_multilingual_v2',
        'voice_settings': {'stability': 0.5, 'similarity_boost': 0.8}
    }
    
    try:
        response = requests.post(url, json=payload, headers=headers, timeout=5)
        response.raise_for_status()
        return response.content
    except requests.exceptions.RequestException as e:
        print(f"ElevenLabs API Error: {e}")
        return None

def pre_generate_audio():
    """Pre-generate audio for all scripted prompts and responses at startup"""
    print("Pre-generating audio files...")
    
    for item in SCRIPT:
        audio = generate_audio(item["text"])
        if audio:
            item["audio_url"] = f"{BASE_URL}/tts_cache?text={urllib.parse.quote(item['text'])}"
    
    for key in RESPONSES:
        audio = generate_audio(RESPONSES[key]["text"])
        if audio:
            RESPONSES[key]["audio_url"] = f"{BASE_URL}/tts_cache?text={urllib.parse.quote(RESPONSES[key]['text'])}"
    
    print("Audio pre-generation complete")

def process_excel_file():
    """Load and process Excel file with lead information"""
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)  
        sheet = wb[SHEET_NAME]
        
        # Ensure columns exist
        if sheet.max_column < 3:
            sheet.cell(row=1, column=3, value="Remarks")
        if sheet.max_column < 4:
            sheet.cell(row=1, column=4, value="Timestamp")
        
        leads = []
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0] or not row[1]:  
                continue
                
            phone = str(row[1]).strip()
            if phone.startswith('='):
                phone = str(row[1])  
            phone = str(phone).replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
            
            # Ensure proper international format for Exotel
            if not phone.startswith('+'):
                if phone.startswith('91') and len(phone) == 12:
                    phone = '+' + phone
                elif phone.startswith('0'):
                    phone = '+91' + phone[1:]
                elif len(phone) == 10:
                    phone = '+91' + phone
            
            lead = {
                "name": row[0],
                "phone": phone,
                "status": row[2] if len(row) > 2 else "",
                "row_num": idx
            }
            leads.append(lead)
            
        return wb, sheet, leads
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        return None, None, []
    
def update_excel_status(wb, sheet, row_num, status, remarks=""):
    """Update Excel with call status and remarks"""
    try:
        # Update status and remarks
        sheet.cell(row=row_num, column=3, value=status)
        sheet.cell(row=row_num, column=4, value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # Apply color based on status
        status_lower = status.lower()
        if status_lower in STATUS_COLORS:
            sheet.cell(row=row_num, column=3).fill = STATUS_COLORS[status_lower]
        
        wb.save(EXCEL_FILE)
        print(f"Updated Excel row {row_num}: {status}")
        return True
    except Exception as e:
        print(f"Error updating Excel: {e}")
        return False

@app.route('/tts_cache')
def tts_cache():
    """Cached TTS endpoint for pre-generated audio"""
    text = request.args.get('text')
    if not text:
        return "Error: no text provided", 400
    
    audio = generate_audio(text)
    if audio:
        return Response(audio, mimetype='audio/mpeg')
    return "Error generating audio", 500

def determine_interest_response(user_input: str) -> tuple:
    """Return both response text and status"""
    user_input = user_input.lower()
    
    interest_keywords = ['interested', 'ஆமா', 'ஆமாம்', 'சரி', 'ok', 'yes', 'details', 'விவரம்', 'வேண்டும்']
    disinterest_keywords = ['not interested', 'இல்லை', 'no', 'later', 'வேண்டாம்', 'தேவையில்லை']
    
    if any(keyword in user_input for keyword in interest_keywords):
        return (RESPONSES["interested"]["audio_url"] or RESPONSES["interested"]["text"], "Interested")
    elif any(keyword in user_input for keyword in disinterest_keywords):
        return (RESPONSES["not interested"]["audio_url"] or RESPONSES["not interested"]["text"], "Not Interested")
    else:
        return (RESPONSES["default"]["audio_url"] or RESPONSES["default"]["text"], "No Response")

@app.route('/tts')
def tts():
    """Text-to-speech endpoint with fallback"""
    text = request.args.get('text')
    if not text:
        return "Error: no text provided", 400
    
    audio = generate_audio(text)
    if audio:
        return Response(audio, mimetype='audio/mpeg')
    return "Error generating audio", 500

@app.route('/start_call', methods=['GET'])
def start_call():
    """Initiate outbound call through Exotel"""
    to = request.args.get('recipient_phone_number')
    phone = request.args.get('phone', '')  # Excel identifier
    
    if not to:
        return "Error: please provide recipient_phone_number", 400

    try:
        # Exotel API endpoint
        url = f"https://{exotel_subdomain}.exotel.in/v1/Accounts/{account_sid}/Calls/connect.json"
        
        auth = (account_sid, auth_token)
        
        payload = {
            "From": exotel_number,
            "To": to,
            "CallerId": exotel_number,
            "Url": f"{BASE_URL}/exotel_callback?phone={phone}",
            "TimeLimit": "60",  # 60 second call time limit
            "StatusCallback": f"{BASE_URL}/call_status?phone={phone}",
            "CallType": "trans"  # For transactional calls
        }
        
        response = requests.post(url, auth=auth, data=payload)
        response.raise_for_status()
        
        return f"Call initiated to {to}. Response: {response.text}"
    except Exception as e:
        print(f"Error initiating call: {str(e)}")
        # Update Excel if call failed
        wb, sheet, _ = process_excel_file()
        if wb and sheet and phone:
            update_excel_status(wb, sheet, int(phone), "Call Failed", str(e))
        return f"Error initiating call: {str(e)}", 500

@app.route('/exotel_callback', methods=['GET', 'POST'])
def exotel_callback():
    """Handle Exotel call flow"""
    phone = request.args.get('phone', '')
    call_sid = request.values.get('CallSid')
    digits = request.values.get('Digits', '')
    
    response = VoiceResponse()
    
    # Initial greeting
    response.play(f"{BASE_URL}/tts?text={urllib.parse.quote(SCRIPT[0]['text'])}")
    
    # Gather input
    gather = Gather(
        input='speech',
        action=f'/exotel_response?stage=1&phone={phone}&call_sid={call_sid}',
        timeout=5,
        speechTimeout='auto',
        language='ta-IN'
    )
    response.append(gather)
    
    return Response(str(response), content_type='application/xml')

@app.route('/exotel_response', methods=['GET', 'POST'])
def exotel_response():
    """Process user response in Exotel flow"""
    stage = int(request.args.get('stage', 1))
    phone = request.args.get('phone', '')
    call_sid = request.args.get('call_sid', '')
    user_response = request.values.get('SpeechResult', '').strip()
    
    response = VoiceResponse()
    
    if stage < len(SCRIPT):
        # Play next script message
        response.play(f"{BASE_URL}/tts?text={urllib.parse.quote(SCRIPT[stage]['text'])}")
        
        # Gather next response
        gather = Gather(
            input='speech',
            action=f'/exotel_response?stage={stage+1}&phone={phone}&call_sid={call_sid}',
            timeout=5,
            speechTimeout='auto',
            language='ta-IN'
        )
        response.append(gather)
    else:
        # Final response handling
        if user_response:
            response_text, status = determine_interest_response(user_response)
            response.play(f"{BASE_URL}/tts?text={urllib.parse.quote(response_text)}")
            
            # Update Excel
            wb, sheet, _ = process_excel_file()
            if wb and sheet and phone:
                update_excel_status(wb, sheet, int(phone), status, user_response)
        
        response.hangup()
    
    return Response(str(response), content_type='application/xml')

@app.route('/call_status', methods=['POST'])
def call_status():
    """Handle Exotel call status updates"""
    phone = request.form.get('phone')
    call_status = request.form.get('CallStatus')
    duration = request.form.get('Duration')
    call_sid = request.form.get('CallSid')
    
    # Update Excel based on call status
    wb, sheet, _ = process_excel_file()
    if wb and sheet and phone:
        if call_status == 'completed':
            # Check if we already updated status during call
            current_status = sheet.cell(row=int(phone), column=3).value
            if not current_status:
                update_excel_status(wb, sheet, int(phone), "No Response", f"Call completed, duration: {duration}")
        elif call_status == 'failed':
            update_excel_status(wb, sheet, int(phone), "Call Failed", f"Call failed, SID: {call_sid}")
    
    return '', 200

def start_flask():
    """Run optimized Flask server"""
    app.run(debug=False, port=5000, use_reloader=False, threaded=True)

if __name__ == '__main__':
    pre_generate_audio()
    
    # Start Flask server
    flask_thread = threading.Thread(target=start_flask)
    flask_thread.daemon = True
    flask_thread.start()
    time.sleep(2)
    
    # Process Excel file and make calls
    wb, sheet, leads = process_excel_file()
    if not leads:
        print("No valid leads found in Excel file")
    else:
        print(f"Found {len(leads)} leads to process")
        
        for lead in leads:
            if lead["status"] and lead["status"].lower() in ["interested", "not interested"]:
                print(f"Skipping {lead['name']} - already processed")
                continue
                
            print(f"Calling {lead['name']} at {lead['phone']}")
            response = requests.get(
                f"http://localhost:5000/start_call",
                params={
                    "recipient_phone_number": lead["phone"],
                    "phone": str(lead["row_num"])  # Pass row number as identifier
                }
            )
            print(response.text)
            
            # Wait between calls
            time.sleep(10)  # Increased delay to avoid rate limits
    
    # Keep main thread alive
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("Shutting down...")
